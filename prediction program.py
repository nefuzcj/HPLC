import os
import pandas as pd
import math
import numpy as np
import random
import time
from openpyxl import Workbook,load_workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

g_fileName = "Input data.xls"
mp0 = []
mp1 = []
g_chromatographicSystem = None
g_solvent = None
g_gradient = None
g_geneticAlgorithm = None
g_population = None

DIAMETER = 0
FLOWU = 0
DEADTIME = 0
DTIME_D = 0
TTUBE = 0
TWIDTH = 0
COLN = 0
PNMAX = 0
TPTN = 0
DAIN = 0
PCROSS = 0
CNUM = 1
PM = 0
MNUM = 8
DAICOUNT = 0

'''
SAMPLEN = sheet.Range("B6").Value
DURINGT = sheet.Range("B24").Value
TPTN = sheet.Range("D24").Value
PNMAX = sheet.Range("A28").Value
DAIN = sheet.Range("B28").Value
DAICOUNT = sheet.Range("C28").Value
PCROSS = sheet.Range("D28").Value
PM = sheet.Range("E28").Value
CNUM = 1
MNUM = 8
MP0Val_b = sheet.Range("A32").Value * 0.01
MP0Val_e = sheet.Range("B32").Value * 0.01
MP0N = sheet.Range("C32").Value
MP1Val_b = sheet.Range("E32").Value * 0.01
MP1Val_e = sheet.Range("F32").Value * 0.01
MP1N = sheet.Range("G32").Value
'''

def readExcel1(fileName):
    df = pd.read_excel(fileName,"OptGradient")
    return df
    
def readParamtersForChromatographicSystem(file):
    ColumnLenth = file.iat[1,0]
    ColumnDiameter = file.iat[1,1]
    Flowrate = file.iat[1,2]
    DeadTime = file.iat[1,3]
    DwellingVolume = file.iat[1,4]
    TubingVolume =file.iat[1,5]
    USPWidthWithoutColumn =file.iat[1,6]
    NumberOfTheoreticalPlates  =file.iat[1,7]
    
    print("ColumnLenth:",ColumnLenth)
    print("ColumnDiameter:",ColumnDiameter)
    print("Flowrate:",Flowrate)
    print("DeadTime:",DeadTime)
    print("DwellingVolume:",DwellingVolume)
    print("TubingVolume:",TubingVolume)
    print("USPWidthWithoutColumn:",USPWidthWithoutColumn)
    print("NumberOfTheoreticalPlates:",NumberOfTheoreticalPlates)


class pt():
    def __init__(self):
        self.t = 0.0
        self.fai = 0.0

class chromosome():
    def __init__(self):
        self.element = [pt(),pt(),pt(),pt(),pt(),pt(),pt(),pt(),pt(),pt(),pt(),pt(),pt(),pt(),pt(),pt()]
        self.maxtr=0.0
        self.minr=0.0
        self.score=0.0
        self.dai=0.0
    def prin(self):
        #print(self.maxtr,self.minr,self.score,self.dai)
        item = self.element
        print("{:.3f} {:.3f} {:.3f} {:.3f}||||{:.3f} {:.3f} {:.3f} {:.3f} {:.3f} {:.3f} {:.3f} {:.3f} {:.3f} {:.3f} {:.3f} {:.3f}".format
        (self.dai,self.maxtr,self.minr,self.score,item[0].t,item[0].fai,item[1].t,item[1].fai,item[2].t,item[2].fai,item[3].t,item[3].fai,
        item[4].t,item[4].fai,item[5].t,item[5].fai))
    
class chromatographicSystem():
    def __init__(self,file):
        self.ColumnLenth = file.iat[1,0]
        self.ColumnDiameter = file.iat[1,1]
        self.Flowrate = file.iat[1,2]
        self.DeadTime = file.iat[1,3]
        self.DwellingVolume = file.iat[1,4]
        self.TubingVolume =file.iat[1,5]
        self.USPWidthWithoutColumn =file.iat[1,6]
        self.NumberOfTheoreticalPlates  =file.iat[1,7]
    def prin(self):
        print("ColumnLenth:",self.ColumnLenth)
        print("ColumnDiameter:",self.ColumnDiameter)
        print("Flowrate:",self.Flowrate)
        print("DeadTime:",self.DeadTime)
        print("DwellingVolume:",self.DwellingVolume)
        print("TubingVolume:",self.TubingVolume)
        print("USPWidthWithoutColumn:",self.USPWidthWithoutColumn)
        print("NumberOfTheoreticalPlates:",self.NumberOfTheoreticalPlates)
        
class solvent():
    def __init__(self,file):
        self.file1 = file.iloc[6:13,0:5]
        self.file1.index = range(0,len(self.file1.index))
        self.file1.columns = [0,1,2,3,4]

    def prin(self):
        print(self.file1)

class gradient():
    def __init__(self,file):
        self.waterContaining = file.iat[22,1]
        self.methanolContaining = file.iat[22,3]
    def prin(self):
        print("waterContaining:",self.waterContaining)
        print("methanolContaining:",self.methanolContaining)

        
def IsinArray(a, x):
    if x in a:
        return True
    else:
        return False

def rndArray(a, lb, ub):
    b = 0
    e = len(a)
    if e > 0:
        a[b] = random.randint(1+lb,ub)
    else:
        a.append(random.randint(1+lb,ub))
    i = b + 1
    while i < e:
        while(1):
            tmp = random.randint(1+ub,lb)
            j = b
            while j < i - 1:
                if math.fabs(tmp - a[j]) < 0.0001:
                    break 
            if j >= i:
                continue
            else:
                break
        a.append(tmp)
        i = i+1
       
def setArray(a, lb, ub, p):
    n = round((ub - lb + 1) * p)
    if n < 1:
        n = 1
    for i in range(0,n):
        a.append(0)
    if math.fabs(p - 1) < 0.000001:# Then   'p=1
        #print("**********",lb)
        i = lb
        while (i < ub+1):
            if len(a) > i-1:
                a[i-1] = i
            else:
                a.append(i)
            i = i+1
    elif p <= 0.5:# Then            'p<=0.5
        rndArray(a, lb, ub)
    else:#                            'p>0.5
        j = 1
        exa = []
        rndArray(exa, lb, ub)
        i = lb
        while i < ub:
            if IsinArray(exa, i) == False:# Then
                a.append(i)
                j = j + 1
            i = i + 1
'''
Public Sub setArray(a() As Long, lb As Long, ub As Long, p As Double)
    Dim b As Long, e As Long, i As Long, j As Long, tmp As Long
    Dim n As Long, exa() As Long
    n = Round((ub - lb + 1) * p)
    If n < 1 Then n = 1
    ReDim a(1 To n) As Long
    If Abs(p - 1) < 0.000001 Then   'p=1
        For i = lb To ub
            a(i) = i
        Next i
    ElseIf p <= 0.5 Then            'p<=0.5
        Call rndArray(a, lb, ub)
    Else                            'p>0.5
        j = 1
        ReDim exa(1 To ub - lb + 1 - n) As Long
        Call rndArray(exa, lb, ub)
        For i = lb To ub
            If IsinArray(exa, i) = False Then
                a(j) = i
                j = j + 1
            End If
        Next i
    End If
End Sub
'''
class geneticAlgorithm():
    def __init__(self,file):
        self.NumOfChromosomes = file.iat[26,0]
        self.MaxGenerations  = file.iat[26,1]
        self.NumGenerationsBestChromosomeSurviving = file.iat[26,2]
        self.CrossoverRate = file.iat[26,3]
        self.MutationRate = file.iat[26,4]
    def prin(self):
        print("NumOfChromosomes:",self.NumOfChromosomes)
        print("MaxGenerations:",self.MaxGenerations)
        print("NumGenerationsBestChromosomeSurviving:",self.NumGenerationsBestChromosomeSurviving)
        print("CrossoverRate:",self.CrossoverRate)
        print("MutationRate:",self.MutationRate)
#'mutation
def mute(m , mpt, dt, dfai):
    chr = chromosome()
    for i in range(0, mpt - 1): #0 1 2
        chr.element[i].t = m.element[i].t
        chr.element[i].fai = m.element[i].fai
    chr.element[mpt-1].t = m.element[mpt-1].t + dt  # 3
    chr.element[mpt-1].fai = m.element[mpt-1].fai + dfai
    '''
    print("*****************mute*************************",dt,dfai)
    m.prin()
    chr.prin()
    print("*****************mute*************************")
    '''
    for i in range(mpt ,TPTN+1):  # 4 5
        chr.element[i].t = m.element[i].t
        chr.element[i].fai = m.element[i].fai
    adjustChromosome(chr,myGradient)
    return chr
#'crossover
def cross(m , f , cpt):
    result = chromosome()
    #print(cpt)
    #m.prin()
    #f.prin()
    for i in range(0,cpt):
        result.element[i].t = m.element[i].t
        result.element[i].fai = m.element[i].fai
        #print(i,"**1111********",result.element[i].t,m.element[i].t,result.element[i].fai,m.element[i].fai)
        #result.prin()
    for i in range(cpt,TPTN+2):
        result.element[i].t = f.element[i].t
        result.element[i].fai = f.element[i].fai
        #print(i,"**222********")
        #result.prin()
        
    return result
    

def processingData(population):
    print("genetic Algorithm processing data!",DAIN)
    PN = len(population)
    preoptchr = chromosome()
    popNew = []
    for k in range(0,DAIN):
        #'crossover
        crossn = []
        setArray(crossn,1,PN*(PN-1),PCROSS)
        icross = 0
        i = 0
        count = PN
        while(i<PN-1):
            j = i + 1
            while(j < PN):
                if IsinArray(crossn, icross):
                    chr1 = population[i]
                    chr2 = population[j]
                    cpt = []
                    setArray(cpt, 1, TPTN - 1, CNUM / (TPTN - 1))
                    ic = 0
                    while ic < len(cpt):
                        crosschr = cross(chr1, chr2, cpt[ic])
                        adjustChromosome(crosschr,myGradient)
                        if IsNewChromosome(population, crosschr, 0, count):
                            count = count + 1
                            crosschr.dai = k+1
                            if len(population)>count:
                                population[count] = crosschr
                            else:
                                population.append(crosschr)
                        crosschr = cross(chr2, chr1, cpt[ic])
                        adjustChromosome(crosschr,myGradient)
                        if IsNewChromosome(population, crosschr, 0, count):
                            count = count + 1
                            crosschr.dai = k+1
                            if len(population)>count:
                                population[count] = crosschr
                            else:
                                population.append(crosschr)
                        ic = ic + 1
                icross += 1
                j = j + 1
            i = i + 1
        
        '''
        print("***************in processingData crossover *************************",k)
        for item in population:
            item.prin()
        print("***************in processingData crossover************************",k)
        '''
        #mutation
        muten = []
        setArray(muten, 1, PN, PM)
        i = 0
        while i < (len(muten)):
            n = muten[i]
            mpt = []
            mtype = []
            setArray(mpt, 1, TPTN - 1, 1 / (TPTN - 1))
            setArray(mtype, 1, 8, MNUM / 8)
            #print("mpt:",mpt)
            #print("mytype:",mtype)
            #exit()
            chr = population[n]
            dt1 = 0.1
            dfai1 = 0.01
            dt2 = 0.1
            dfai2 = 0.01
            j = 0
            mutechr = chromosome()
            while j < (len(mpt)):
                if mtype[j] == 1:
                    mutechr = mute(chr, mpt[j], -1 * dt1, -1 * dfai1)
                elif mtype[j] == 2:        
                    mutechr = mute(chr, mpt[j], -1 * dt1, 0)
                elif mtype[j] == 3:        
                    mutechr = mute(chr, mpt[j], -1 * dt1, dfai2)
                elif mtype[j] == 4:        
                    mutechr = mute(chr, mpt[j], 0, -1 * dfai1)
                elif mtype[j] == 5:
                    mutechr = mute(chr, mpt[j], 0, dfai2)
                elif mtype[j] == 6:
                    mutechr = mute(chr, mpt[j], dt2, -1 * dfai1)
                elif mtype[j] == 7:
                    mutechr = mute(chr, mpt[j], dt2, 0)
                elif mtype[j] == 8:
                    mutechr = mute(chr, mpt[j], dt2, dfai2)
                adjustChromosome(mutechr,myGradient)
                if IsNewChromosome(population, mutechr, 0, count):
                    count = count + 1
                    '''
                    chr.prin()
                    print("***** 7777 **",len(mtype),mpt[j],len(population),count)
                    mutechr.prin()
                    '''
                    if len(population)>count:
                        mutechr.dai = population[count].dai
                        population[count] = mutechr
                    else:
                        mutechr.dai = k
                        population.append(mutechr)
                j += 1
            i += 1
        '''
        print("***************in processingData*************************")
        for item in population:
            item.prin()
        print("***************in processingData************************")
        '''
        
        #print("*****111******",count)
        for i in range(0, len(population)):
            score(population[i])
        #print("*****222******",count)
        population.sort(key=takeSecond1,reverse=True)
        optchr = population[0]
        #'exits the loop
        '''
        print("***************in processingData*************************")
        for item in population:
            item.prin()
        print("***************in processingData************************")
        '''
        if k == 1:
            preoptchr = optchr
            DAICOUNT = 1
        else:
            if IsEqual(optchr, preoptchr):
                DAICOUNT = DAICOUNT + 1
            else:
                DAICOUNT = 1
                print("k ****Generating data*11111***")
                preoptchr = optchr
        if DAICOUNT >= 20:
            break
        k += 1
        #print("DAICOUNT",DAICOUNT,k)
        #break       # add for test
    #print("*****444******",count)

class population():
    def __init__(self,file):
        self.MinMobilePhase = file.iat[30,0]*0.01
        self.MaxMobilePhase = file.iat[30,1]*0.01
        self.NumMobilePhase = file.iat[30,2]
        self.IntervalMobilePhase = file.iat[30,3]
        self.MinEndMobilePhase = file.iat[30,4]*0.01
        self.MaxEndMobilePhase = file.iat[30,5]*0.01
        self.NumEndMobilePhase = file.iat[30,6]
        self.IntervalEndMobilePhase = file.iat[30,7]        
    def prin(self):
        print("MinMobilePhase:",self.MinMobilePhase)
        print("MaxMobilePhase:",self.MaxMobilePhase)
        print("NumMobilePhase:",self.NumMobilePhase)
        print("IntervalMobilePhase:",self.IntervalMobilePhase)
        print("MinEndMobilePhase:",self.MinEndMobilePhase)
        print("MaxEndMobilePhase:",self.MaxEndMobilePhase)
        print("NumEndMobilePhase:",self.NumEndMobilePhase)
        print("IntervalEndMobilePhase:",self.IntervalEndMobilePhase)

def adjustChromosome(a,gradient):
    for i in range(1,TPTN):
        if a.element[i].t < a.element[i - 1].t:
            a.element[i].t = a.element[i - 1].t + 0.01
            print("**222**********")
            a.prin()
            exit()
        if a.element[i].fai > g_population.MaxEndMobilePhase:
            a.element[i].fai = g_population.MaxEndMobilePhase
        if a.element[i].fai < a.element[i - 1].fai:
            a.element[i].fai = a.element[i - 1].fai 
def IsEqual(a,b):
    result = True
    for i in range(0,g_gradient.methanolContaining,1):
        t_temp = math.fabs(a.element[i].t - b.element[i].t)
        fai_temp = math.fabs(a.element[i].fai - b.element[i].fai)
        if math.fabs(a.element[i].t - b.element[i].t) > 0.0001 or math.fabs(a.element[i].fai - b.element[i].fai) > 0.00001:
            result = False
            break
    return result;
'''
'judge if two chromosomes are the same
Public Function IsEqual(a As chromosome, b As chromosome) As Boolean
    Dim i As Long, t As Boolean
    t = True
    For i = 0 To TPTN
        If Abs(a.element(i).t - b.element(i).t) > 0.0001 Or _
            Abs(a.element(i).fai - b.element(i).fai) > 0.00001 Then
            t = False
            Exit For
        End If
    Next i
    IsEqual = t
End Function
'''
def IsNewChromosome(p,c,b,e):
    result = True
    i = b
    for i in range(0,e,1):
        if IsEqual(p[i], c):
            result = False
            break
    return result;
'''
'judge if the chromosome is new
Public Function IsNewChromosome(p() As chromosome, c As chromosome, b As Long, e As Long) As Boolean
    Dim i As Long, a As Boolean
    a = True
    For i = b To e
        If IsEqual(p(i), c) Then
            a = False
            Exit For
        End If
    Next i
    IsNewChromosome = a
End Function
'''         

#make time program for mobile phase composition dimensionless
def AssignProgramDim4chr(dimprogram, c):
    for i in range(0,g_gradient.methanolContaining+1):
        dimprogram[i][0] = c.element[i].t / g_chromatographicSystem.DeadTime
        dimprogram[i][1] = c.element[i].fai
        #print(i,dimprogram[i][0],dimprogram[i][1],c.element[i].t,c.element[i].fai)
    
#calculate score
def calScore(minr, maxtr):
    n = 5
    result = 0
    DURINGT = g_gradient.waterContaining
    if DURINGT - maxtr < 0:
        result = 0
    else:
        if minr > 9.99:
            s2 = 9.99 * 10 ** n
        else:
            s2 = round(minr, 2) * 10 ** n
        dt = DURINGT - maxtr
        if dt > 99.9:
            s3 = 999
        else:
            s3 = round(dt, 1) * 10
        result = s2 + s3
    return result


#time program for mobile phase composition
def FaiProgram(dimt, dimprogram):
    fai = 0
    
    for stage in range(0,len(dimprogram) - 2):
        tb = dimprogram[stage][0]
        te = dimprogram[stage + 1][0]
        if dimt > tb and dimt <= te:
            break
    if stage < len(dimprogram)-1:
        if (te - tb) / (tb + 0.001) < 0.000000001:# 'stepwise
            fai = dimprogram[stage][1]
        else:#    'linear
            faib = dimprogram[stage][1]
            faie = dimprogram[stage+1][1]
            slope = (faie - faib) / (te - tb)
            fai = faib + slope * (dimt - tb)
    else:
        fai = -1
    if fai == 0:
        print("stage:",stage,"dimt:",dimt,"dimprg:",dimprogram,len(dimprogram))
        
    return fai

#calculate mobile phase composition, fai
def calFai(dimt, dimx , dimtd, dimprogram):
    f0 = dimprogram[0][1]
    result = 0.0
    if dimt <= dimx + dimtd:
        result = f0
    else:
        result = FaiProgram(dimt - dimx - dimtd, dimprogram)
    return result


#'for Integral_gauss4Xdim procedure
def f(fai, sample):
    return 1 / calk_nonLSS(fai, sample)

def calXIntegral(s, fai0, fai1):
    result = 0.0
    Pi = math.pi
    s2 = s.iloc[0,2+2]
    s1 = s.iloc[0,1+2] * -1
    k0 = math.exp(s.iloc[0,0+2])
    Lower_limit = math.sqrt(s2) * fai0 + s1 / 2 / math.sqrt(s2)
    Upper_limit = math.sqrt(s2) * fai1 + s1 / 2 / math.sqrt(s2)
    t1 = math.erf(Upper_limit) - math.erf(Lower_limit)
    result = math.sqrt(Pi) / 2 / k0 / math.sqrt(s2) * math.exp(s1 ** 2 / 4 / s2) * t1
    return result
    
#Gaussian quadrature for retention time
def Integral_gauss4Xdim(lb, ub, sample):
    result = 0.0
    t = [-0.9061798459, 
                    -0.5384693101,
                    0, 
                    0.5384693101, 
                    0.9061798459]
    aa = [0.2369268851, 
                    0.4786286705, 
                    0.5688888889, 
                    0.4786286705, 
                    0.2369268851]
    tol = 0.00001
    n = 10
    Is1st = True
    while(1):
        s = 0
        h = (ub - lb) / n
        for i in range(1, n,1):
            l1 = lb + (i - 1) * h
            l2 = lb + i * h
            for j in range(0,4):
                x = (l2 - l1) / 2 * t[j] + (l2 + l1) / 2
                ff = f(x, sample)
                s = s + (l2 - l1) * aa[j] * ff / 2
        if Is1st:
            v1 = s
            Is1st = False
        else:
            v2 = s
            if math.fabs(v2 - v1) / (math.fabs(v2) + 1) < tol:
                break
            else:
                v1 = v2
        n = n + 10
    return v2
    
#calculate the distance of solute migrating in the column within the gradient segment
def calX_dim(dimprogram, sample, td_dim, tinj_dim, x_dim):
    #print("calX_dim",)
    n = len(dimprogram)
    n1 = sample.size-3
    fai0 = dimprogram[0][1]
    kk0 = calk_nonLSS(fai0, sample)
    x_dim[0] = (td_dim - tinj_dim) / kk0
    for stage in range(1,n):
        t0 = dimprogram[stage - 1][0]
        fai0 = dimprogram[stage - 1][1]
        t1 = dimprogram[stage][0]
        fai1 = dimprogram[stage][1]
        if t1 - t0 < 0:
            print("The next time is less than the current time!",t1,t0,stage)
            break
        elif (t1 - t0) / (t0 + 0.001) < 0.000000001: # Then 'stepwise
            x_dim[stage] = x_dim[stage - 1]
        elif math.fabs(fai1 - fai0) / (fai0 + 0.001) < 0.000000001:# Then 'isocratic
            kk0 = calk_nonLSS(fai0, sample)
            x_dim[stage] = x_dim[stage - 1] + (t1 - t0) / kk0
        else:# 'linear
            bk = (fai1 - fai0) / (t1 - t0)
            if (n1 == 1):
                #LSSM
                kk0 = calk_nonLSS(fai0, sample)
                kk1 = calk_nonLSS(fai1, sample)
                s = sample.iloc[0,2]
                x_dim[stage] = x_dim[stage - 1] + 1 / bk / s * (1 / kk1 - 1 / kk0)
                #'QSSM
            elif n1 == 2:
                x_dim[stage] = x_dim[stage - 1] + calXIntegral(sample, fai0, fai1) / bk
                #'others
            else:
                x_dim[stage] = x_dim[stage - 1] + Integral_gauss4Xdim(fai0, fai1, sample) / bk

#Judge where the solute is eluted in the gradient profile
def EluteN(x_dim):
    result = -1
    for i in range(0,len(x_dim),1):
        if x_dim[i] >= 1:
            result = i
            break
    return result
def calf(fai, fai0, x0, dimslope, sample):
    return Integral_gauss4Xdim(fai0, fai, sample) / dimslope + x0 - 1


#calculate mobile phase composition at which the analyte exits the column, faiR
def calfaiR_gauss(fai0, x0, dimslope, sample):
    k = calk_nonLSS(fai0, sample)
    x = fai0 - dimslope * k * (x0 - 1)
    while(1):
        if x < 0 or x > 1:
            x = Rnd
        f = calf(x, fai0, x0, dimslope, sample)
        if math.fabs(f) < 0.00000001:
            break
        else:
            k = calk_nonLSS(x, sample)
            x = x - dimslope * k * f
    return x

#calculate inverse error function
def invErf(val):
    result = 0.0
    sign = True
    if math.fabs(val) > 1:
        print( "invErf Invalid")
        exit()
    if val < 0:
        val = -val
        sign = False
    x1 = 0
    x2 = 5
    while(calf2(x2, val) < 0):
        x2 = 1.2*x2
    f1 = calf2(x1, val)
    f2 = calf2(x2, val)
    while(1):
        x = (x1 * f2 - x2 * f1) / (f2 - f1)
        f = calf2(x, val)
        if math.fabs(f) <= 0.000001:
            v = x
            break
        else:
            if f * f1 > 0:
                x1 = x
            else:
                x2 = x
    if sign:
        result = v 
    else:
        result = -1 * v
    return result
#For invErf procedure
def calf2(x,val):
    result = 0.0
    result = math.erf(x) - val
    return result


#calculate faiR for QSSM
def calfaiR2(s, slope, fai0, x0):
    result = 0.0
    Pi = math.pi
    s2 = s.iloc[0,2+2]
    s1 = s.iloc[0,1+2] * -1
    k0 = math.exp(s.iloc[0,0+2])
    
    Lower_limit = math.sqrt(s2) * fai0 + s1 / 2 / math.sqrt(s2)
    t1 = math.erf(Lower_limit)
    
    t2 = 2 * slope * k0 * math.sqrt(s2) * (1 - x0) / math.sqrt(Pi) * math.exp(-1 * s1 ** 2 / 4 / s2)
    result = 1 / math.sqrt(s2) * invErf(t1 + t2) - s1 / 2 / s2
    return result
    
    
#caculate retention time
def calRT_dim(dimprogram, line, td_dim, tinj_dim):
    result = 0.0
    #print("calRT_dim",td_dim, tinj_dim)
    n = len(dimprogram)
    n1 = line.size - 3
    x = []
    for i in range(0,n):
        x.append(0)
    
    calX_dim(dimprogram, line, td_dim, tinj_dim, x)
    n = EluteN(x)
    #when the solute cannot be eluted from the column
    if n < 0:
        result = 999 / DEADTIME
        return result
    #when the solute is eluted in dwelling time
    if n == 0:
        fai0 = dimprogram[0,1]
        kk0 = calk_nonLSS(fai0, line)
        result = 1 + kk0
        return result
    t0 = dimprogram[n - 1][0]
    t1 = dimprogram[n][0]
    fai0 = dimprogram[n - 1][1]
    fai1 = dimprogram[n][1]
    bn = (fai1 - fai0) / (t1 - t0)
    kk = calk_nonLSS(fai0, line)
    x0 = x[n - 1]
    if math.fabs(bn) < 0.000000001:# Then   'isocratic
        tr = 1 + t0 + td_dim + kk * (1 - x0)
    else:
        if n1 == 1:
            s = line.iloc[0,2]
            tr = 1 + t0 + td_dim + 1 / bn / s * Log(1 + bn * s * kk * (1 - x0))
        elif n1 == 2:#'QSSM
            faiR = calfaiR2(line, bn, fai0, x0)
            tr = 1 + t0 + td_dim + (faiR - fai0) / bn
        else:
            faiR = calfaiR_gauss(fai0, x0, bn, line)
            tr = 1 + t0 + td_dim + (faiR - fai0) / bn
    result = tr 
    return result
    
    
#solvent strength model
def calk_nonLSS(fai, sample):
    result = 0.0
    n = sample.size - 2
    s = 0
    for i in range(0,n):
        if i == 1:
            coe = -1
        else:
            coe = 1
        s = s + coe * sample.iloc[0,i+2] * fai ** i
    try:
        result = math.exp(s)
    except(OverflowError):
        print("calk_nonLSS",s)
    return result
    
#For the calculation of G
def sumGIntegral(dimprogram, dimtr, sample):
    Pi = math.pi
    s2 = sample.iloc[0,4]
    s1 = sample.iloc[0,3] * -1
    k0 = math.exp(sample.iloc[0,2])
    n1 = sample.size - 3
    t = dimtr * DEADTIME - DEADTIME - DTIME_D
    if t > dimprogram[len(dimprogram)-1][0] * DEADTIME:# Then
        print("The solute cannot be eluted within the gradient program!")
        exit()
    for stage in range(1,len(dimprogram),1):
        tb = dimprogram[stage - 1][0] * DEADTIME
        te = dimprogram[stage][0] * DEADTIME
        if t > tb and t <= te:
            break
    #[1,N-1]
    s = 0
    for i in range(1,stage,1):
        tb = dimprogram[i - 1][0] * DEADTIME
        te = dimprogram[i][0] * DEADTIME
        #not stepwise
        if math.fabs(te - tb) > 0.000001:
            faib = dimprogram[i - 1][1]
            faie = dimprogram[i][1]
            slope = (faie - faib) / (te - tb)
            if slope < 0.000001: #isocratic
                kb = calk_nonLSS(faib, sample)
                s = s + (te - tb) * (1 + kb) ** 2 / kb ** 3
                if s < 0:
                    print("sum 111:",s,te,tb,kb)
                    exit()
            else:#'linear
                if n1 == 1:# 'LSSM
                    kb = calk_nonLSS(faib, sample)
                    ke = calk_nonLSS(faie, sample)
                    s = s + 1 / slope / (-1 * s1) * (1 / 3 * (1 / ke ** 3 - 1 / kb ** 3) + (1 / ke ** 2 - 1 / kb ** 2) + (1 / ke - 1 / kb))
                    if s < 0:
                        print("sum 222:",s,slope,s1,ke,kb)
                        exit()
                elif n1 == 2:#'QSSM
                    Lower_limit = math.sqrt(s2) * faib + s1 / 2 / math.sqrt(s2)
                    Upper_limit = math.sqrt(s2) * faie + s1 / 2 / math.sqrt(s2)
                    temp = math.erf(Upper_limit) - math.erf(Lower_limit)

                    t1 = math.sqrt(Pi) / 2 / slope / k0 / math.sqrt(s2) * math.exp(s1 ** 2 / 4 / s2) * temp
                    Lower_limit = math.sqrt(2 * s2) * faib + s1 / math.sqrt(2 * s2)
                    Upper_limit = math.sqrt(2 * s2) * faie + s1 / math.sqrt(2 * s2)
                    temp = math.erf(Upper_limit) - math.erf(Lower_limit)

                    t2 = math.sqrt(Pi) / slope / k0 ** 2 / math.sqrt(2 * s2) * math.exp(s1 ** 2 / 2 / s2) * temp
                    Lower_limit = math.sqrt(3 * s2) * faib + math.sqrt(3) * s1 / 2 / math.sqrt(s2)
                    Upper_limit = math.sqrt(3 * s2) * faie + math.sqrt(3) * s1 / 2 / math.sqrt(s2)
                    temp = math.erf(Upper_limit) - math.erf(Lower_limit)
                    t3 = math.sqrt(Pi) / 2 / slope / k0 ** 3 / math.sqrt(3 * s2) * math.exp(3 * s1 ** 2 / 4 / s2) * temp
                    s = s + t1 + t2 + t3
                    if s < 0:
                        print("sum 333:",s,t1,t2,t3)
                        exit()
                    #print("sumGIntegral s1:",s)
    # [N-1,N]
    tb = dimprogram[stage - 1][0] * DEADTIME
    te = dimprogram[stage][0] * DEADTIME
    faib = dimprogram[stage - 1][1]
    faie = dimprogram[stage][1]
    slope = (faie - faib) / (te - tb)
    if slope < 0.00000001:# 'isocratic
        kb = calk_nonLSS(faib, sample)
        s = s + (dimtr * DEADTIME - DEADTIME - DTIME_D - tb) * (1 + kb) ** 2 / kb ** 3
        if s < 0:
            print("sum 444:",s,dimtr,DEADTIME,DTIME_D,tb,kb)
            exit()
    else:#'linear
        faiR = FaiProgram((dimtr * DEADTIME - DEADTIME - DTIME_D) / DEADTIME, dimprogram)
        if n1 == 1: #LSSM
            kb = calk_nonLSS(faib, sample)
            ke = calk_nonLSS(faiR, sample)
            s = s + 1 / slope / (-1 * s1) * (1 / 3 * (1 / ke ** 3 - 1 / kb ** 3) + (1 / ke ** 2 - 1 / kb ** 2) + (1 / ke - 1 / kb))
        elif n1 == 2:#'QSSM
            Lower_limit = math.sqrt(s2) * faib + s1 / 2 / math.sqrt(s2)
            Upper_limit = math.sqrt(s2) * faiR + s1 / 2 / math.sqrt(s2)
            temp = math.erf(Upper_limit) - math.erf(Lower_limit)
            t1 = math.sqrt(Pi) / 2 / slope / k0 / math.sqrt(s2) * math.exp(s1 ** 2 / 4 / s2) * temp
            if t1 < 0:
                print("sum 444**** 11111 :",t1,Lower_limit,Upper_limit,temp,faib,faiR)
            Lower_limit = math.sqrt(2 * s2) * faib + s1 / math.sqrt(2 * s2)
            Upper_limit = math.sqrt(2 * s2) * faiR + s1 / math.sqrt(2 * s2)
            temp = math.erf(Upper_limit) - math.erf(Lower_limit)
            t2 = math.sqrt(Pi) / slope / k0 ** 2 / math.sqrt(2 * s2) * math.exp(s1 ** 2 / 2 / s2) * temp
            Lower_limit = math.sqrt(3 * s2) * faib + math.sqrt(3) * s1 / 2 / math.sqrt(s2)
            Upper_limit = math.sqrt(3 * s2) * faiR + math.sqrt(3) * s1 / 2 / math.sqrt(s2)
            temp = math.erf(Upper_limit) - math.erf(Lower_limit)
            t3 = math.sqrt(Pi) / 2 / slope / k0 ** 3 / math.sqrt(3 * s2) * math.exp(3 * s1 ** 2 / 4 / s2) * temp
            s = s + t1 + t2 + t3
            if s < 0:
                print("sum 555:",s,t1,t2,t3)
                exit()
            #print("sumGIntegral s2:",s)
    #print("sumGIntegral s3:",s)
    return s
def calGfromSS(dimprogram, tr, sample):
    #print("calGfromSS")
    result = 0.0
    if tr <= DEADTIME + DTIME_D:
        result = 1
    else:
        faiR = calFai(tr / DEADTIME, 1, DTIME_D / DEADTIME, dimprogram)
        kfaiR = calk_nonLSS(faiR, sample)
        fai0 = dimprogram[0][1]
        kfai0 = calk_nonLSS(fai0, sample)
        t1 = DTIME_D * (1 + kfai0) ** 2 / kfai0 ** 3
        t2 = sumGIntegral(dimprogram, tr / DEADTIME, sample)
        #print("calGfromSS:",kfaiR,t1,t2)
        try:
            result = math.sqrt(kfaiR ** 2 / DEADTIME / (1 + kfaiR) ** 2 * (t1 + t2))
        except(IndexError,ValueError):
            print("calGfromSS:",kfaiR,DEADTIME,t1,t2,kfaiR ** 2 / DEADTIME / (1 + kfaiR) ** 2 * (t1 + t2))
    return result;

def takeSecond(elem):
    return elem[0]
#score for chromosome
def score(c):
    #print("score ************0000000000 ")
    dimprogram = []
    trsigma = []
    for i in range(0,g_gradient.methanolContaining+1):
        temp= [0.0,0.0]
        dimprogram.append(temp)
    AssignProgramDim4chr(dimprogram,c)
    lineNum = g_solvent.file1.iloc[:,0].size - 1 #获取mySolvent行数
    for i in range(0,lineNum):
        temp=[0.0,0.0]
        temp[0] = calRT_dim(dimprogram, g_solvent.file1.iloc[[i+1]], DTIME_D / DEADTIME, 0) * DEADTIME
        if math.fabs(temp[0] - 999) < 0.000001:
            temp[1] = 999
        else:
            kfaiR = calk_nonLSS(calFai(temp[0] / DEADTIME, 1, DTIME_D / DEADTIME, dimprogram), g_solvent.file1.iloc[[i+1]])
            GVal = calGfromSS(dimprogram, temp[0], g_solvent.file1.iloc[[i+1]])
            temp[1] = GVal * DEADTIME * (1 + kfaiR) / math.sqrt(COLN)
            #print("score:",kfaiR,GVal,temp[1])
        #print("***************1111******************",i)
        #print("score ",temp)           
        trsigma.append(temp)
        #break #test add muty
    trsigma.sort(key=takeSecond)
    
    #Call ascendsort(trsigma, 0)  升序排序trsigma[][0]
    #for item in trsigma:
    #   print(item[0],item[1])
    r = []
    #print("score ************22222222222 ")
    for i in range(0,lineNum-1):
        tr1 = trsigma[i][0]
        tr2 = trsigma[i+1][0]
        sigma1 = trsigma[i][1]
        sigma2 = trsigma[i + 1][1]
        
        #print(tr1,tr2,sigma1,sigma2,r)
        r.append((tr2 - tr1) / 2 / (sigma1 + sigma2))
    c.maxtr = round(trsigma[lineNum-1][0], 3)
    c.minr = round(min(r), 3)
    c.score = round(calScore(c.minr, c.maxtr),0)
    #print("score ************333333333333 ")
    #print(c.maxtr,c.minr,c.score)

def takeSecond1(elem):
    return elem.score    
def initData(population,gradient):
    d = (g_population.MaxMobilePhase - g_population.MinMobilePhase) / (g_population.NumMobilePhase - 1)
    d = round(d,3)
    #print(d,g_population.MaxMobilePhase,g_population.MinMobilePhase,g_population.NumMobilePhase)
    for i in range(1,g_population.NumMobilePhase+1,1):
        tem = g_population.MinMobilePhase + (i - 1) * d
        tem = round(tem,3)
        mp0.append(tem)
    
    d = (g_population.MaxEndMobilePhase - g_population.MinEndMobilePhase) / (g_population.NumEndMobilePhase - 1)
    d = round(d,3)
    #print(d,g_population.MaxEndMobilePhase,g_population.MinEndMobilePhase,g_population.NumEndMobilePhase)
    for i in range(1,g_population.NumEndMobilePhase+1,1):
        tem = g_population.MinEndMobilePhase + (i - 1) * d
        tem = round(tem,3)
        mp1.append(tem)
        
    dt = gradient.waterContaining / gradient.methanolContaining
    count = 0
    temp = []
    #print(mp0,mp1)
    for i in range(1,g_population.NumEndMobilePhase+1,1):
        #print(g_population.NumMobilePhase)
        for j in range(1,g_population.NumMobilePhase+1,1):
            chr = chromosome()
            chr.element[0].t = 0
            chr.element[0].fai = mp0[j-1]
            chr.element[gradient.methanolContaining].t = gradient.waterContaining
            chr.element[gradient.methanolContaining].fai = mp1[i-1]
            for k in range(1,(gradient.methanolContaining),1):
                fai0 = chr.element[0].fai
                fai1 = chr.element[gradient.methanolContaining].fai
                t = k * dt
                #print(dt,t)
                fai = fai0 + (fai1 - fai0) / gradient.waterContaining * t
                chr.element[k].t = round(t,3)
                chr.element[k].fai = round(fai,3)
            #print(chr.element[0].t,chr.element[0].fai)
            #chr.prin()
            adjustChromosome(chr,gradient)
            #print("****************",i, j,"******************")
            #chr.prin()

            chr.dai = 0
            if count == 0:
                count = count + 1
                temp.append(chr)
                #print("********",chr,count)
                #chr.prin()
            else:
                #print("********",chr,count)
                #chr.prin()
                if IsNewChromosome(temp, chr, 0, count):
                    count = count + 1
                    temp.append(chr)
    
    #for item in temp:
    #    print(item.element[0].t,item.element[0].fai)
    for item in range(0,count):
        score(temp[item])
        #temp[item].prin()
        #exit()
        
    temp.sort(key=takeSecond1,reverse=True)
    if count >= PNMAX:
        for i in range(0,PNMAX):
            population.append(temp[i])
        count = PNMAX
    else:
        for i in range(0,count):
            population.append(temp[i])
    PN = count

#make time program for mobile phase composition dimensionless
def AssignProgramDim4chr(dimprogram, c):
    i = 0 
    while i < TPTN+1:
        dimprogram[i][0] = c.element[i].t / DEADTIME
        dimprogram[i][1] = c.element[i].fai
        #print("dimprogram:",dimprogram[i][0],dimprogram[i][1])
        i += 1

#calculate retention time and peak width
def calTrsigma(trsigma, dimprogram):
    lineNum = g_solvent.file1.iloc[:,0].size - 1
    for i in range(0,lineNum):
        tem = [0,0,0,0]
        tem[0] = calRT_dim(dimprogram,g_solvent.file1.iloc[[i+1]], DTIME_D / DEADTIME, 0) * DEADTIME
        kfaiR = calk_nonLSS(calFai(tem[0] / DEADTIME, 1, DTIME_D / DEADTIME, dimprogram), g_solvent.file1.iloc[[i+1]])
        GVal = calGfromSS(dimprogram, tem[0], g_solvent.file1.iloc[[i+1]])
        tem[1] = GVal * DEADTIME * (1 + kfaiR) / math.sqrt(COLN)
        tem[2] = g_solvent.file1.iloc[[i+1]].iloc[0,0]
        tem[3] = GVal
        #print("calTrsigma:",tem[0],tem[1],tem[2],tem[3]);
        trsigma.append(tem)
        i += 1
def setchart(g_fileName):
    wb = load_workbook("Output.xlsx")
    ws = wb["ResultOpt"]
    maxRow = ws.max_row
    chart = ScatterChart()
    analyte = ScatterChart()
    Chromatogram = ScatterChart()
    
    chart.title = "Gradient"
    analyte.title = "Chromatogram for each analyte"
    Chromatogram.title = "Chromatogram in sum"

    Chromatogram.style = analyte.style = chart.style = 5
    chart.y_axis.title = 'Time (min)'
    
    x = Reference(ws, min_col=1, min_row=1, max_row=6)
    y = Reference(ws, min_col=2, min_row=1, max_row=6)
    s = Series(y, xvalues=x)
    chart.series.append(s)
    
    x = Reference(ws, min_col=6, min_row=1, max_row=maxRow)
    '''
    y = Reference(ws, min_col=7, min_row=1, max_row=maxRow)
    s = Series(y, xvalues=x)
    analyte.series.append(s)
    y = Reference(ws, min_col=8, min_row=1, max_row=maxRow)
    s = Series(y, xvalues=x)
    analyte.series.append(s)
    y = Reference(ws, min_col=9, min_row=1, max_row=maxRow)
    s = Series(y, xvalues=x)
    analyte.series.append(s)
    y = Reference(ws, min_col=10, min_row=1, max_row=maxRow)
    s = Series(y, xvalues=x)
    analyte.series.append(s)
    '''
    y = Reference(ws, min_col=13, min_row=1, max_row=maxRow)
    s = Series(y, xvalues=x)
    analyte.series.append(s)
    '''
    x = Reference(ws, min_col=1, min_row=1, max_row=6)
    y = Reference(ws, min_col=2, min_row=1, max_row=6)
    s = Series(y, xvalues=x)
    Chromatogram.series.append(s)
    '''
    ws.add_chart(chart, "L5")
    ws.add_chart(analyte, "L20")
    ws.add_chart(Chromatogram, "L40")
    wb.save("Output.xlsx")
#correct USP width
def AdjUSPWidth(w , exw ):
    s = math.sqrt((w / 4) ** 2 + (exw / 4) ** 2)
    return s * 4
def _normfun(x, mu, sigma):
    pdf = np.exp(-((x - mu)**2)/(2*sigma**2)) / (sigma * np.sqrt(2*np.pi))
    return pdf

#score format
def FormatScore(score, n, dn):
    print("FormatScore")
    
'''
    Dim s1 As String, s2 As String, n1 As Long, i As Long
    s1 = Trim(Str(score))
    n1 = Len(s1)
    If n1 < n Then
        For i = 1 To n - n1
            s1 = "0" & s1
        Next i
    End If
    s2 = ""
    For i = 1 To Len(s1) - dn Step dn
        s2 = Mid(s2, 1, Len(s2)) & Mid(s1, i, dn) & ","
    Next i
    FormatScore = Mid(s2, 1, Len(s2)) & Mid(s1, Len(s1) - dn + 1, Len(s1))  
'''
#output results
def chromatogram(trsigma, dimprogram,fileName):
    df = pd.ExcelWriter("Output.xlsx",engine='openpyxl',mode='w')
    sheet = pd.DataFrame()
    lb = 0
    ub = len(trsigma)
    i = lb
    while i < ub:
        temp = [0,0,0,0,0]
        tr = trsigma[i][0]
        sigma = trsigma[i][1]
        temp[0] = (tr + TTUBE)   #'RT
        temp[1] = AdjUSPWidth(4 * sigma, TWIDTH) / 4 #'USP width
        temp[2] = trsigma[i][2]      #'No.
        temp[3] = calFai(tr / DEADTIME, 1, DTIME_D / DEADTIME, dimprogram)   #'faiR
        temp[4] = trsigma[i][3]   #'G
        trsigma[i] = temp
        i += 1
    soluten = ub - lb + 1
    for item in trsigma:
        print("in chromatogram trsigma:",item[0],item[1],item[2],item[3])
    r = []
    #ascendsort(trsigma, 0) #排序
    trsigma.sort(key=takeSecond,reverse=False)
    soluten = len(trsigma)
    tend = trsigma[ub-1][0]
    print("in chromatogram tend:",tend,"soluten:",soluten)
    i = 0
    sheet[0]='1'
    sheet[1]=''
    sheet[2]=''
    sheet[3]=''
    sheet[4]=''
    sheet[5]=''
    sheet[6]=''
    sheet[7]=''
    sheet[8]=''
    sheet[9]=''
    sheet[10]=''
    sheet[11]=''
    sheet[12]=''
    sheet[13]=''
    sheet[14]=''
    index_X = 0 #行
    index_Y = 0 #列
    #print(DEADTIME)
    while i < len(dimprogram):
        sheet.loc[index_X] = ''
        sheet.iloc[index_X, 0] = dimprogram[i][0] * DEADTIME
        sheet.iloc[index_X, 1] = dimprogram[i][1] * 100
        i += 1
        index_X += 1
    i = 0
    
    while i < soluten - 1:
        tr1 = trsigma[i][0]
        tr2 = trsigma[i + 1][0]
        w1 = 4 * trsigma[i][1]
        w2 = 4 * trsigma[i + 1][1]
        r.append(2 * (tr2 - tr1) / (w1 + w2))
        i += 1
    maxtr = round(trsigma[soluten-1][0], 3)
    minr = round(min(r), 2)
    score1 = calScore(minr, maxtr)
    row = 9
    for i in range(sheet.iloc[:,0].size,sheet.iloc[:,0].size+30):
        #print("add sheet line 1",i)
        sheet.loc[i] = ''
    sheet.iloc[row, 0] = "Score="
    sheet.iloc[row, 1] = '{0:,}'.format(score1)#FormatScore(score1, 6, 3)
    sheet.iloc[row + 1, 0] = "minr=" 
    sheet.iloc[row + 1, 1] = minr
    sheet.iloc[row + 2, 0] = "maxtr="
    sheet.iloc[row + 2, 1] = maxtr
    sheet.iloc[row + 4, 0] = "Sample No."
    sheet.iloc[row + 4, 1] = "tR (min)"
    sheet.iloc[row + 4, 2] = "USP Width (min)"
    sheet.iloc[row + 4, 3] = "R"
    sheet.iloc[row + 4, 4] = "faiR"
    j = row + 5
    i = 0
    while i < soluten:
        print(i)
        n = trsigma[i][2]
        sheet.iloc[j, 0] = "{0:0>2d}".format(n)+"\\"+"{0:s}".format(mySolvent.file1.iloc[n,1])#StrNo(n, 2) & "/" & (FindName(s, n))
        sheet.iloc[j, 1] = round(trsigma[i][0], 3)
        sheet.iloc[j, 2] = round(4 * trsigma[i][1], 5)
        if i >= 1:
            sheet.iloc[j, 3] = round(r[i - 1], 2)
        sheet.iloc[j, 4] = round(trsigma[i][3] * 100, 2)
        j += 1
        i += 1
    #'G
    j = j + 5
    sheet.iloc[j, 0] = "Sample No."
    sheet.iloc[j, 1] = "G"
    j = j + 1
    
    for i in range(0,soluten):
        n = trsigma[i][2]
        sheet.iloc[j, 0] = "{0:0>2d}".format(n)+"\\"+"{0:s}".format(mySolvent.file1.iloc[n,1])#StrNo(n, 2) & "/" & (FindName(s, n))
        #tr = trsigma[i][0]
        #sigma = trsigma[i][1]
        sheet.iloc[j, 1] = trsigma[i][3]
        j = j + 1
        i += 1
    
    cb = 6
    i = 0
    while(1):
        if sheet.iloc[:,0].size < i+1:
            sheet.loc[i] = ''
        ss = 0
        t = (i) * 0.01
        #
        try:
            sheet.iloc[i, cb-1] = t
        except(IndexError):
            print(sheet)
            print("err i cb :",i,cb,sheet.iloc[:,0].size)
        j = cb 
        while j < cb + soluten:
            n = j - cb
            tr = trsigma[n][0]
            sigma = trsigma[n][1]
            try:
                text = _normfun(t,tr,sigma)
                sheet.iloc[i,j] = _normfun(t,tr,sigma)
            except(IndexError):
                print("err",i,j,text)
            ss = ss + float(sheet.iloc[i, j])
            j += 1 
        sheet.iloc[i, j] = ss
        #print("data:",sheet.iloc[i,5],sheet.iloc[i,6],sheet.iloc[i,7],sheet.iloc[i,8],sheet.iloc[i,9],sheet.iloc[i,10])
        '''
        if i > 2000:
            sheet.to_excel(df,"ResultOpt",index=False,header=False)
            df.save()
            print("exit")        
            exit()
        '''
        i = i + 1
        if t >= tend + 3:
            break
    sheet.iloc[8,0] = "Generations"
    sheet.iloc[8,1] = optchr.dai
    #print(sheet)
    sheet.to_excel(df,"ResultOpt",index=False,header=False)
    df.save()
if __name__ == "__main__":
    timeStart = time.time()
    file = readExcel1(g_fileName)
    
    myChromatographicSystem = chromatographicSystem(file)
    mySolvent = solvent(file)
    mySolvent.prin()
    #lineNum = mySolvent.file1.iloc[:,0].size - 1  #获取mySolvent行数  4
    #line = mySolvent.file1.iloc[[1]]              # 5
    #print(lineNum,line.size,line,line.iloc[0,1])
    #exit()
    myGradient = gradient(file)
    myGeneticAlgorithm = geneticAlgorithm(file)
    myPopulation = population(file)
    '''mySolvent.prin()
    
    myGradient = gradient(file)
    myGradient.prin()
    
    myGeneticAlgorithm = geneticAlgorithm(file)
    myGeneticAlgorithm.prin()
    myPopulation = population(file)
    myPopulation.prin()
    '''
    g_chromatographicSystem = myChromatographicSystem
    g_solvent = mySolvent
    g_gradient = myGradient
    g_geneticAlgorithm = myGeneticAlgorithm
    g_population = myPopulation
    COLL = g_chromatographicSystem.ColumnLenth
    DIAMETER = g_chromatographicSystem.ColumnDiameter
    FLOWU = g_chromatographicSystem.Flowrate
    DEADTIME = g_chromatographicSystem.DeadTime
    DTIME_D = g_chromatographicSystem.DwellingVolume / FLOWU
    TTUBE = g_chromatographicSystem.TubingVolume * 0.001 / FLOWU
    TWIDTH = g_chromatographicSystem.USPWidthWithoutColumn
    COLN = g_chromatographicSystem.NumberOfTheoreticalPlates
    PNMAX = myGeneticAlgorithm.NumOfChromosomes
    DAIN = myGeneticAlgorithm.MaxGenerations
    PCROSS = myGeneticAlgorithm.CrossoverRate
    PM = myGeneticAlgorithm.MutationRate
    TPTN = g_gradient.methanolContaining
    DAICOUNT = 0
    myPop = []
    
    initData(myPop,myGradient)
    
    print("***************init end*************************")
    for item in myPop:
        item.prin()
    print("***************init end************************")
    
    processingData(myPop)
    
    print("***************processingData*************************")
    for item in myPop:
        item.prin()
    print("***************processingData*************************")
    
    dimprogram = []
    for i in range(0,g_gradient.methanolContaining+1):
        temp= [0.0,0.0]
        dimprogram.append(temp)
    
    
    optchr = myPop[0]
    
    trsigma = []
    optchr.prin()
    #根据optchr，获取dimprogram
    AssignProgramDim4chr(dimprogram, optchr)
    print(len(dimprogram))
    for i in dimprogram:
        print("in main AssignProgramDim4chr end",i[0],i[1])
    #根据dimprogram，获取trsigma
    calTrsigma(trsigma,dimprogram)
    i = 0 
    print(len(trsigma))
    for item in trsigma:
        print("in main calTrsigma end",item[0],item[1],item[2],item[3])
    #写xlsx
    chromatogram(trsigma, dimprogram,g_fileName)
    #向xlsx中绘图
    setchart(g_fileName)
    #结束时间
    timeEnd = time.time()

    #运行时间
    timeCount = timeEnd - timeStart

    print("Total elapsed time(seconds):",timeCount)
    
    wb = load_workbook("Output.xlsx")
    ws = wb["ResultOpt"]
    ws.cell(row=1, column=4,value='Start time:')
    ws.cell(row=1, column=5,value=time.asctime(time.localtime(timeStart)))
    ws.cell(row=2, column=4,value='End time:')
    ws.cell(row=2, column=5,value=time.asctime(time.localtime(timeEnd)))
    ws.cell(row=3, column=4,value='Total elapsed time(seconds):')
    ws.cell(row=3, column=5,value=timeEnd - timeStart)
    wb.save("Output.xlsx")
